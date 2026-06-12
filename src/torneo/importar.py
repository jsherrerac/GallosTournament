"""Importación masiva desde Excel (plantilla del cliente).
Columnas esperadas (fila de encabezado): CUERDA, FRENTE, ALIANZA, CIUDAD, PESO, COLOR, TIPO, MARCA, PLACA, ANILLO"""
import openpyxl

COLS = ["CUERDA", "FRENTE", "ALIANZA", "CIUDAD", "PESO", "COLOR", "TIPO", "MARCA", "PLACA", "ANILLO"]

def leer_excel(ruta):
    """Devuelve (filas, errores). Cada fila es un dict con las columnas."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    # encontrar la fila de encabezados (la que contenga 'CUERDA' y 'PLACA')
    encabezado, fila_inicio = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(v).strip().upper().rstrip("*").strip() if v else "" for v in row]
        if "CUERDA" in vals and "PLACA" in vals:
            encabezado, fila_inicio = vals, i + 1
            break
    if encabezado is None:
        return [], ["No se encontró la fila de encabezados (CUERDA, PLACA...)."]
    idx = {c: encabezado.index(c) for c in COLS if c in encabezado}
    filas, errores = [], []
    for n, row in enumerate(ws.iter_rows(min_row=fila_inicio, values_only=True), fila_inicio):
        if not row or all(v is None or str(v).strip()=="" for v in row):
            continue
        d = {c: (row[i] if i < len(row) else None) for c, i in idx.items()}
        if not d.get("CUERDA") or not d.get("PLACA") or d.get("PESO") is None:
            errores.append(f"Fila {n}: falta CUERDA, PLACA o PESO — ignorada.")
            continue
        filas.append(d)
    return filas, errores

def importar(torneo, ruta):
    """Carga el Excel al torneo. Devuelve (n_gallos, errores)."""
    filas, errores = leer_excel(ruta)
    n = 0
    for d in filas:
        nombre = str(d["CUERDA"]).strip()
        if torneo.buscar_cuerda(nombre) is None:
            torneo.registrar_cuerda(nombre, str(d.get("ALIANZA") or ""),
                                    str(d.get("CIUDAD") or ""))
        try:
            frente = d.get("FRENTE") or 1
            frente = int(str(frente).upper().replace("F", "").strip() or 1)
            torneo.registrar_gallo(
                nombre, frente, d["PLACA"], d["PESO"],
                color=str(d.get("COLOR") or ""), tipo=str(d.get("TIPO") or "GALLO"),
                marca=d.get("MARCA") or "", anillo=d.get("ANILLO") or "",
                ciudad=str(d.get("CIUDAD") or ""))
            n += 1
        except ValueError as e:
            errores.append(f"Placa {d.get('PLACA')}: {e}")
    return n, errores

def crear_plantilla(ruta):
    """Genera el Excel de plantilla para repartir a las cuerdas."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "Inscripción"
    ws.append(["PLANTILLA DE INSCRIPCIÓN — TORNEO GALLÍSTICO"])
    ws.append(["Cada fila = un ave. PESO en libras-onzas (ej: 3-4 = 3 libras 4 onzas)."])
    ws.append(COLS)
    ws.append(["GUALCALA", 1, "", "Centro", "3-2", "Morado", "POLLO", "06", "767270", "345"])
    wb.save(ruta)
